import streamlit as st
import pandas as pd
import io
from datetime import datetime, timedelta
from database import (
    init_db, authenticate_user, get_time_codes, log_attendance, 
    get_user_attendance, apply_leave, get_my_leaves, get_team_members, 
    get_pending_team_leaves, update_leave_status_with_remarks, get_users_by_role, 
    get_all_attendance_logs, users_col, update_password,
    get_attendance_for_date, get_team_leave_history, log_attendance_from_leave
)
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

# --- INITIALIZATION ---
init_db()
st.set_page_config(page_title="LMATS", layout="wide")

if "logged_in" not in st.session_state:
    st.session_state.logged_in = False
    st.session_state.user = None

def logout():
    st.session_state.logged_in = False
    st.session_state.user = None
    st.rerun()

def get_manager_options():
    managers = get_users_by_role("Manager")
    return {m.get("user_id"): m.get("username", "Unknown") for m in managers if isinstance(m, dict)}

# --- LOGIN / SIGN UP ---
if not st.session_state.logged_in:
    st.title("LMATS Portal")
    tab1, tab2 = st.tabs(["Login", "Create Account"])
    with tab1:
        with st.form("login_form"):
            username = st.text_input("Username")
            password = st.text_input("Password", type="password")
            if st.form_submit_button("Login"):
                user = authenticate_user(username, password)
                if user:
                    st.session_state.logged_in = True
                    st.session_state.user = user
                    st.rerun()
                else: st.error("Invalid credentials.")
    with tab2:
        with st.form("signup_form"):
            new_uid, new_uname = st.text_input("User ID"), st.text_input("Username")
            new_pwd = st.text_input("Password", type="password")
            new_role = st.selectbox("Role", ["Employee", "Manager"]) 
            mgrs = get_manager_options()
            selected_mgr = st.selectbox("Assign Manager", options=list(mgrs.keys()), format_func=lambda x: mgrs[x])
            if st.form_submit_button("Sign Up"):
                users_col.insert_one({"user_id": new_uid, "username": new_uname, "password": new_pwd, "role": new_role, "manager_id": selected_mgr})
                st.success("Account created!")
    st.stop()

# --- SIDEBAR & SESSION ---
user_data = st.session_state.user
role = user_data.get("role")
user_id = user_data.get("user_id")
st.sidebar.title(f"Welcome, {user_data.get('username', 'User')}")
st.sidebar.button("Logout", on_click=logout)

# Contextual Navigation based on Roles
if role == "Admin":
    # Admin does NOT have Apply Leave or Bulk Mark Attendance
    nav = ["Team Dashboard", "Daily Dashboard", "Proxy Logging", "System Config", "Compliance Export (TSDATA)", "Settings"]
elif role == "Manager":
    nav = ["My Dashboard", "Team Dashboard", "Daily Dashboard", "Bulk Mark Attendance", "Apply Leave", "Proxy Logging", "Settings"]
else:
    nav = ["My Dashboard", "Daily Dashboard", "Bulk Mark Attendance", "Apply Leave", "Settings"]

selection = st.sidebar.radio("Navigation", nav)

# --- MODULES ---
if selection == "My Dashboard":
    st.header("My Dashboard")
    logs = get_user_attendance(user_id)
    if logs: 
        st.dataframe(pd.DataFrame(logs)[["Date", "TimeCode", "Place", "Location", "is_locked"]], use_container_width=True)
    
    st.subheader("My Leave Status")
    leaves = get_my_leaves(user_id)
    if leaves: 
        df = pd.DataFrame(leaves)
        for col in ["StartDate", "EndDate", "Type", "Status", "ManagerRemarks"]:
            if col not in df.columns:
                df[col] = ""
        st.table(df[["StartDate", "EndDate", "Type", "Status", "ManagerRemarks"]])

elif selection == "Daily Dashboard":
    st.header("Daily Attendance Status")
    logs = get_attendance_for_date(datetime.today().strftime("%Y-%m-%d"))
    present_ids = [l.get("UserID") for l in logs if isinstance(l, dict)]
    
    users = get_users_by_role() if role in ["Manager", "Admin"] else [user_data]
    users = [u for u in users if isinstance(u, dict)]
    
    c1, c2 = st.columns(2)
    with c1:
        st.success("Present")
        for u in users:
            if u.get("user_id") in present_ids: st.write(f"✅ {u.get('username', 'Unknown')}")
    with c2:
        st.error("Absent")
        for u in users:
            if u.get("user_id") not in present_ids: st.write(f"❌ {u.get('username', 'Unknown')}")

elif selection == "Bulk Mark Attendance":
    st.header("Bulk Mark Attendance")
    with st.form("att_form"):
        dates = st.date_input("Date Range", value=(datetime.today(), datetime.today()))
        code = st.selectbox("Time Legend", get_time_codes())
        if st.form_submit_button("Log"):
            start = dates[0] if isinstance(dates, tuple) else dates
            end = dates[1] if isinstance(dates, tuple) and len(dates) > 1 else start
            for i in range((end - start).days + 1):
                cur = start + timedelta(days=i)
                if cur.weekday() < 5: log_attendance(user_id, cur, code, "N/A", "N/A", "", user_id, False)
            st.success("Successfully logged.")

elif selection == "Apply Leave":
    st.header("Apply Leave")
    with st.form("leave_form"):
        s, e = st.date_input("Date Range", value=(datetime.today(), datetime.today()))
        l_type = st.selectbox("Type", ["Sick", "Casual", "Earned", "Unpaid"])
        reason = st.text_area("Reason")
        if st.form_submit_button("Submit"):
            s_d = s if not isinstance(s, tuple) else s[0]
            e_d = e if not isinstance(e, tuple) else (e[1] if len(e)>1 else e[0])
            apply_leave(user_id, user_data.get("manager_id"), s_d, e_d, l_type, reason)
            st.success("Leave applied.")

elif selection == "Settings":
    st.header("Settings")
    with st.form("pwd"):
        o = st.text_input("Old Password", type="password")
        n = st.text_input("New Password", type="password")
        c = st.text_input("Confirm Password", type="password")
        if st.form_submit_button("Update Password"): 
            if n == c:
                update_password(user_id, n)
                st.success("Password successfully updated.")
            else:
                st.error("New passwords do not match.")

elif selection == "Team Dashboard" and role in ["Manager", "Admin"]:
    st.header("Team Management")
    
    # Section: Pending Leaves
    st.subheader("Pending Leave Requests")
    pending_leaves = get_pending_team_leaves(user_id)
    if not pending_leaves:
        st.info("No pending leave requests at the moment.")
    else:
        for leave in pending_leaves:
            if not isinstance(leave, dict): continue
            u_info = users_col.find_one({"user_id": leave.get('UserID')})
            name = u_info.get('username', 'Unknown') if u_info else "Unknown"
            with st.expander(f"Request: {name} ({leave.get('UserID')}) - {leave.get('Type')} from {leave.get('StartDate')} to {leave.get('EndDate')}"):
                new_type = st.selectbox("Assign Type", get_time_codes(), key=f"t_{leave['_id']}")
                remarks = st.text_input("Remarks", key=f"r_{leave['_id']}")
                if st.button("Approve", key=f"a_{leave['_id']}"):
                    update_leave_status_with_remarks(leave["_id"], "Approved", remarks)
                    log_attendance_from_leave(leave.get('UserID'), leave.get('StartDate'), leave.get('EndDate'), new_type, user_id)
                    st.rerun()
                if st.button("Reject", key=f"re_{leave['_id']}"):
                    update_leave_status_with_remarks(leave["_id"], "Rejected", remarks)
                    st.rerun()

    # Section: Leave History
    st.subheader("Leave Request History")
    history_leaves = get_team_leave_history(user_id)
    if not history_leaves:
        st.info("No approved or rejected leave history.")
    else:
        history_data = []
        for h in history_leaves:
            if not isinstance(h, dict): continue
            u_info = users_col.find_one({"user_id": h.get('UserID')})
            name = u_info.get('username', 'Unknown') if u_info else "Unknown"
            history_data.append({
                "Employee": f"{name} ({h.get('UserID')})",
                "Start Date": h.get("StartDate"),
                "End Date": h.get("EndDate"),
                "Type": h.get("Type"),
                "Status": h.get("Status"),
                "Manager Remarks": h.get("ManagerRemarks", "")
            })
        st.table(pd.DataFrame(history_data))

elif selection == "Proxy Logging" and role in ["Manager", "Admin"]:
    st.header("Proxy Logging")
    u_list = get_users_by_role() if role == "Admin" else get_team_members(user_id)
    u_dict = {u.get("user_id"): u.get("username", "Unknown") for u in u_list if isinstance(u, dict) and u.get("user_id")}
    tid = st.selectbox("Select Employee", list(u_dict.keys()), format_func=lambda x: f"{u_dict[x]} ({x})")
    with st.form("proxy"):
        dates = st.date_input("Date Range", value=(datetime.today(), datetime.today()))
        c = st.selectbox("Code", get_time_codes())
        if st.form_submit_button("Log"):
            start = dates[0] if isinstance(dates, tuple) else dates
            end = dates[1] if isinstance(dates, tuple) and len(dates) > 1 else start
            for i in range((end - start).days + 1):
                cur = start + timedelta(days=i)
                if cur.weekday() < 5: log_attendance(tid, cur, c, "Proxy", "N/A", "", user_id, True)
            st.success(f"Logged for {u_dict[tid]}")

elif selection == "System Config" and role == "Admin":
    st.header("System Config")
    with st.form("add_user_form"):
        new_uid = st.text_input("User ID")
        new_uname = st.text_input("Username")
        new_pwd = st.text_input("Password", type="password")
        new_role = st.selectbox("Role", ["Employee", "Manager", "Admin"])
        
        mgrs = get_manager_options()
        mgr_options = [None] + list(mgrs.keys())
        def format_mgr(x): return "None" if x is None else mgrs.get(x, x)
        selected_mgr = st.selectbox("Assign Manager (Optional)", options=mgr_options, format_func=format_mgr)
        
        if st.form_submit_button("Create User"):
            if users_col.find_one({"user_id": new_uid}):
                st.error("User ID already exists!")
            elif users_col.find_one({"username": new_uname}):
                st.error("Username already exists!")
            else:
                users_col.insert_one({
                    "user_id": new_uid, 
                    "username": new_uname, 
                    "password": new_pwd, 
                    "role": new_role, 
                    "manager_id": selected_mgr
                })
                st.success(f"User '{new_uname}' added successfully.")

elif selection == "Compliance Export (TSDATA)" and role == "Admin":
    st.header("Annual TSDATA Matrix")
    
    current_year = datetime.today().year
    fy_options = [f"FY {y}-{str(y+1)[-2:]}" for y in range(2020, 2035)]
    default_fy = f"FY {current_year}-{str(current_year+1)[-2:]}"
    default_idx = fy_options.index(default_fy) if default_fy in fy_options else 6
    
    fy_selection = st.selectbox("Select Financial Year", fy_options, index=default_idx)
    
    if st.button("Generate Matrix"):
        fy_start_year = int(fy_selection[3:7])
        s_date = pd.Timestamp(year=fy_start_year, month=4, day=1)
        e_date = pd.Timestamp(year=fy_start_year+1, month=3, day=31)
        
        all_days = pd.date_range(start=s_date, end=e_date)
        logs = get_all_attendance_logs()
        
        all_users = get_users_by_role()
        target_users = sorted(
            [u for u in all_users if isinstance(u, dict) and u.get('role') in ['Employee', 'Manager']], 
            key=lambda x: x.get('username', '')
        )
        
        wb = Workbook()
        ws = wb.active
        ws.title = fy_selection.replace(" ", "_")
        
        # Styles for Header
        header_font = Font(bold=True)
        # Using a standard light blue fill for the headers
        blue_fill = PatternFill(start_color="9BC2E6", end_color="9BC2E6", fill_type="solid")
        
        # EXACT CSV MATCH: Row 1 & 2 & 3 Headers for Col 1 and 2
        
        # Cell A1
        c_a1 = ws.cell(row=1, column=1, value="TIMEDATA")
        c_a1.font = header_font
        c_a1.fill = blue_fill
        ws.cell(row=1, column=2).fill = blue_fill # Fill empty neighbor for uniform look
        
        # Cell A2 / B2
        c_a2 = ws.cell(row=2, column=1, value="DATE")
        c_a2.font = header_font
        c_a2.fill = blue_fill
        
        c_b2 = ws.cell(row=2, column=2, value="DAY")
        c_b2.font = header_font
        c_b2.fill = blue_fill
        
        # Row 3 fills for purely visual alignment
        ws.cell(row=3, column=1).fill = blue_fill
        ws.cell(row=3, column=2).fill = blue_fill

        # Dynamic Columns Start at Col 3 (C)
        col = 3
        for u in target_users:
            uname = u.get('username', 'Unknown')
            uid = u.get('user_id', 'Unknown')
            
            # Row 1: Username
            c_uname = ws.cell(row=1, column=col, value=uname)
            c_uname.font = header_font
            c_uname.fill = blue_fill
            ws.cell(row=1, column=col+1).fill = blue_fill
            ws.cell(row=1, column=col+2).fill = blue_fill
            
            # Row 2: User ID
            c_uid = ws.cell(row=2, column=col, value=uid)
            c_uid.font = header_font
            c_uid.fill = blue_fill
            ws.cell(row=2, column=col+1).fill = blue_fill
            ws.cell(row=2, column=col+2).fill = blue_fill
            
            # Row 3: Detail Headers (Changed 'Code' to 'Time')
            c_t = ws.cell(row=3, column=col, value="Time")
            c_t.font = header_font
            c_t.fill = blue_fill
            
            c_p = ws.cell(row=3, column=col+1, value="Place")
            c_p.font = header_font
            c_p.fill = blue_fill
            
            c_da = ws.cell(row=3, column=col+2, value="DA")
            c_da.font = header_font
            c_da.fill = blue_fill
            
            col += 3
            
        log_dict = {(pd.to_datetime(l['Date']).strftime('%Y-%m-%d'), l.get('UserID')): l for l in logs if isinstance(l, dict)}
        
        row = 4
        for day in all_days:
            # Row 4+: Dates in Col A
            ws.cell(row=row, column=1, value=day.strftime('%d/%b/%Y'))
            # Row 4+: Day in Col B
            ws.cell(row=row, column=2, value=day.strftime('%a')) # Will output Mon, Tue, etc.
            
            # Row 4+: Values start in Col C
            col = 3
            for u in target_users:
                rec = log_dict.get((day.strftime('%Y-%m-%d'), u.get('user_id')))
                if rec:
                    ws.cell(row=row, column=col, value=rec.get('TimeCode', ''))
                    ws.cell(row=row, column=col+1, value=rec.get('Place', ''))
                    ws.cell(row=row, column=col+2, value=rec.get('DA', ''))
                col += 3
            row += 1
            
        # Freeze panes and column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 6
        ws.freeze_panes = 'C4'
        
        output = io.BytesIO()
        wb.save(output)
        
        file_name_export = f"TSDATA_{fy_selection.replace(' ', '_')}.xlsx"
        st.download_button("📥 Download Excel Report", data=output.getvalue(), file_name=file_name_export)